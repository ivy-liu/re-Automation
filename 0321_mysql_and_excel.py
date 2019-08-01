import mysql.connector

# 连接数据库
try:
    con = mysql.connector.connect(
        host='127.0.0.1',
        user='root',
        password='vertrigo',
        port=3306,
        database='xiaomingtest',
        charset='utf8',
        buffered=True
    )
    print('数据库连接成功')
    cursor = con.cursor()  # 获取游标
except mysql.connector.Error as e:
    print('数据库连接失败', str(e))


# 创建student表  auto_increment primary key主键自增，创建数据时不需要再写入
Student = '''
    create table Student(
        StdID int auto_increment primary key,
        StdName varchar(100) not null,
        Gender enum('M','F') not null,
        Age int
    )
'''

try:
    cursor.execute(Student)
    print('创建表成功')
except Exception as e:

    print('创建表失败', str(e))
# finally:
#     cursor.close()
#     print('关闭数据库连接')

# 插入数据

try:
    # 字符串插入
    print('----字符串插入---')

    stu_info = '''
    insert into Student(StdName,Gender,Age)
    values('霸王龙','M',27)
    '''
    cursor.execute(stu_info)

    # 元组插入
    print('----元组插入---')
    # 此处%s为占位符，不是格式化字符串，所以age用%s
    stu_info_tup = '''
    insert into student(StdName,Gender,Age)values(%s,%s,%s)
    '''
    data = ('小强', 'M', 21)
    cursor.execute(stu_info_tup, data)

    # 提交操作，不提交不生效，2选1
    con.commit()
    # cursor.execute('commit')

    print('插入数据成功')
except Exception as e:
    print('插入失败', str(e))

# 查询数据
stu_select = '''
    select *from Student where StdID>150
    '''
try:
    cursor.execute(stu_select)
    # 获取所有记录列表 
    # fetchall 抓取全部 
    # fetchmany(2)返回2条数据 
    # fetchone() result[0],result[1],result[2]得到StdID,StdName,Gender...
    results = cursor.fetchall()
    for row in results:
        StdID = row[0]
        StdName = row[1]
        Gender = row[2]
        Age = row[3]
    
        # 打印结果
        print ("StdID=%s,StdName=%s,Gender=%s,Age=%s"% 
        (StdID, StdName, Gender, Age))
except Exception as e:
    print('查询失败', str(e))



cursor.close()  # 关闭游标
con.close()  # 关闭连接



'''
1.连接数据库
2.创建游标
3.执行cursor.execute(stu_info)，‘stu_info’SQL语句
4.cursor.close()  # 关闭游标  con.close()  # 关闭连接
'''


print('\n-----------excel------------\n')
# import openpyxl
from openpyxl.styles import Font,colors,alignment
from openpyxl import Workbook
# load_workbook已存在 Workbook不存在可创建,如存在就

import time,os

# os.chdir('D:\\python_code\\re_Automation')#切换目录至

path='D:\\python_code\\re_Automation\\excel_demo.xlsx'
#上方路径切换过了可以直接传值名字，否则传完整路径

#创建excel文件，操作结束后需要wb.save(path)
wb=Workbook()
# wb=load_workbook(path)
# 报错zipfile.BadZipFile: File is not a zip file，还没有找到解决方法。。。
ws=wb.create_sheet('testsheet00')
wb.create_sheet('testsheet01')
wb.create_sheet('testsheet02')
print('ws是啥',ws)#ws是啥 <WriteOnlyWorksheet "testsheet01">
ws1=wb.active
print('ws1是啥',ws1)#ws1是啥 <Worksheet "Sheet">
ws2=wb['testsheet01']
print('ws2是啥',ws2)#ws2是啥 <Worksheet "testsheet">

# ss=wb.get_sheet_names()#get_sheet_names已弃用 用wb.sheetnames
# print('获取工作簿中所有工作表名',ss)
# s1=wb.get_sheet_names()[0]#get_sheet_names已弃用 用wb.sheetnames
# print('方法1：获取某个工作表名',s1)
# s2=wb.get_sheet_by_name('sheet1')#get_sheet_by_name已弃用，用wb[sheetname]
# print('方法2：获取某个工作表名的对象，供后续调用其它方法使用',s2)

print('工作簿中所有工作表名',wb.sheetnames) #返回列表


#设置单元格数值
ws['A1']=123.11
ws['B1']='你好'
ws['B2']='ws你好呀，我是小花'
ws1['a1']='ws1haha'
ws2['b3']='ws2哈哈哈'

#获取单元格数据
print('录入单元格A1值',ws['A1'].value)
print('录入单元格B1值',ws1['a1'].value)
print('录入单元格B2值',ws2['B3'].value)

#save the file
wb.save(path)

#切换表格
# ws=wb.create_sheet('testsheet01')
# ws1=wb.active
# ws2=wb['testsheet']
